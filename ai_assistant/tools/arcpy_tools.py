"""Actual ArcGIS Pro operations, executed against the CURRENT open project.

Every function here is a leaf action dispatched by tools/registry.py. Keep
return values small and JSON-serializable -- they go straight back into the
model's context window.
"""
import re

import arcpy


def _current_map():
    aprx = arcpy.mp.ArcGISProject("CURRENT")
    m = aprx.activeMap
    if m is None:
        raise RuntimeError("No active map. Open a map view in ArcGIS Pro first.")
    return aprx, m


def _find_layer(map_obj, layer_name):
    for lyr in map_obj.listLayers():
        if lyr.name == layer_name:
            return lyr
    raise ValueError(f"No layer named '{layer_name}' found in the active map.")


def list_layers():
    _, m = _current_map()
    result = []
    for lyr in m.listLayers():
        entry = {"name": lyr.name, "is_feature_layer": bool(lyr.isFeatureLayer)}
        if lyr.supports("DATASOURCE"):
            entry["source"] = lyr.dataSource
        if lyr.isFeatureLayer:
            desc = arcpy.Describe(lyr)
            entry["geometry_type"] = getattr(desc, "shapeType", None)
        result.append(entry)
    return {"layers": result}


def describe_layer(layer_name):
    _, m = _current_map()
    lyr = _find_layer(m, layer_name)
    desc = arcpy.Describe(lyr)
    fields = [{"name": f.name, "type": f.type, "alias": f.aliasName} for f in desc.fields]
    count = int(arcpy.management.GetCount(lyr)[0])
    extent = getattr(desc, "extent", None)
    return {
        "name": lyr.name,
        "geometry_type": getattr(desc, "shapeType", None),
        "spatial_reference": desc.spatialReference.name if desc.spatialReference else None,
        "feature_count": count,
        "fields": fields,
        "extent": (
            {"xmin": extent.XMin, "ymin": extent.YMin, "xmax": extent.XMax, "ymax": extent.YMax}
            if extent
            else None
        ),
    }


def query_features(layer_name, where_clause, fields=None, max_records=50):
    _, m = _current_map()
    lyr = _find_layer(m, layer_name)
    desc = arcpy.Describe(lyr)
    field_names = fields or [f.name for f in desc.fields if f.type not in ("Geometry", "Blob", "Raster")]
    rows = []
    truncated = False
    with arcpy.da.SearchCursor(lyr, field_names, where_clause=where_clause) as cursor:
        for i, row in enumerate(cursor):
            if i >= max_records:
                truncated = True
                break
            rows.append(dict(zip(field_names, [str(v) for v in row])))
    return {"fields": field_names, "rows": rows, "truncated": truncated}


def select_by_attribute(layer_name, where_clause, selection_type="NEW_SELECTION"):
    _, m = _current_map()
    lyr = _find_layer(m, layer_name)
    arcpy.management.SelectLayerByAttribute(lyr, selection_type, where_clause)
    count = 0 if selection_type == "CLEAR_SELECTION" else int(arcpy.management.GetCount(lyr)[0])
    return {"selected_count": count}


def buffer_layer(layer_name, distance, out_name, distance_unit="Feet"):
    aprx, m = _current_map()
    lyr = _find_layer(m, layer_name)
    out_path = f"{aprx.defaultGeodatabase}\\{out_name}"
    arcpy.analysis.Buffer(lyr, out_path, f"{distance} {distance_unit}")
    m.addDataFromPath(out_path)
    return {"created": out_path}


def create_feature_class(name, geometry_type, spatial_reference_wkid, fields=None):
    aprx, m = _current_map()
    sr = arcpy.SpatialReference(spatial_reference_wkid)
    arcpy.management.CreateFeatureclass(aprx.defaultGeodatabase, name, geometry_type, spatial_reference=sr)
    out_path = f"{aprx.defaultGeodatabase}\\{name}"
    for field in fields or []:
        length = field.get("length", 255) if field["type"] == "TEXT" else None
        arcpy.management.AddField(out_path, field["name"], field["type"], field_length=length)
    m.addDataFromPath(out_path)
    return {"created": out_path}


def add_features(layer_name, features):
    _, m = _current_map()
    lyr = _find_layer(m, layer_name)
    attr_field_names = sorted({k for f in features for k in (f.get("attributes") or {})})
    field_names = ["SHAPE@JSON"] + attr_field_names
    inserted = 0
    with arcpy.da.InsertCursor(lyr, field_names) as cursor:
        for f in features:
            geom = arcpy.AsShape(f["geometry"], False)  # False = GeoJSON-style input
            attrs = f.get("attributes") or {}
            cursor.insertRow([geom.JSON] + [attrs.get(name) for name in attr_field_names])
            inserted += 1
    return {"inserted": inserted}


def add_data_to_map(path):
    _, m = _current_map()
    m.addDataFromPath(path)
    return {"added": path}


def run_geoprocessing_tool(toolbox_alias, tool_name, parameters):
    toolbox = getattr(arcpy, toolbox_alias, None)
    if toolbox is None:
        raise ValueError(f"Unknown toolbox alias '{toolbox_alias}'.")
    tool = getattr(toolbox, tool_name, None)
    if tool is None:
        raise ValueError(f"No tool '{tool_name}' in toolbox '{toolbox_alias}'.")
    result = tool(**parameters)
    messages = list(result.getMessages()) if hasattr(result, "getMessages") else [str(result)]
    return {"messages": messages}


# ---------------------------------------------------------------------------
# Project / map / layout control (arcpy.mp) -- everything geoprocessing tools
# don't reach: which maps and layouts exist, layer visibility, camera/zoom,
# bookmarks, and exporting a map or layout to an image/PDF.
# ---------------------------------------------------------------------------

def _current_project():
    return arcpy.mp.ArcGISProject("CURRENT")


def _find_map(aprx, map_name):
    for m in aprx.listMaps():
        if m.name == map_name:
            return m
    raise ValueError(f"No map named '{map_name}' found in the project.")


def _find_layout(aprx, layout_name):
    for lyt in aprx.listLayouts():
        if lyt.name == layout_name:
            return lyt
    raise ValueError(f"No layout named '{layout_name}' found in the project.")


def list_maps():
    aprx = _current_project()
    active = aprx.activeMap.name if aprx.activeMap else None
    return {"maps": [m.name for m in aprx.listMaps()], "active_map": active}


def add_map(name, template="Map"):
    aprx = _current_project()
    m = aprx.createMap(name, template)
    return {"created_map": m.name}


def set_active_map(map_name):
    aprx = _current_project()
    m = _find_map(aprx, map_name)
    m.openView()  # opening a map view makes it the active view/map in this session
    return {"active_map": m.name}


def set_layer_visibility(layer_name, visible):
    _, m = _current_map()
    lyr = _find_layer(m, layer_name)
    lyr.visible = bool(visible)
    return {"layer": lyr.name, "visible": lyr.visible}


def zoom_to_layer(layer_name):
    aprx, m = _current_map()
    lyr = _find_layer(m, layer_name)
    view = aprx.activeView
    if not hasattr(view, "camera"):
        raise RuntimeError(
            "The active view isn't a map view (it may be a layout). Open/activate "
            "a map view in ArcGIS Pro first."
        )
    extent = arcpy.Describe(lyr).extent
    view.camera.setExtent(extent)
    return {"zoomed_to": lyr.name}


def list_bookmarks():
    aprx = _current_project()
    m = aprx.activeMap
    if m is None:
        raise RuntimeError("No active map. Open a map view in ArcGIS Pro first.")
    return {"bookmarks": [b.name for b in m.listBookmarks()]}


def apply_bookmark(bookmark_name):
    aprx = _current_project()
    m = aprx.activeMap
    if m is None:
        raise RuntimeError("No active map. Open a map view in ArcGIS Pro first.")
    matches = [b for b in m.listBookmarks() if b.name == bookmark_name]
    if not matches:
        raise ValueError(f"No bookmark named '{bookmark_name}' found in the active map.")
    view = aprx.activeView
    if not hasattr(view, "camera"):
        raise RuntimeError("The active view isn't a map view. Open/activate a map view first.")
    view.camera.setExtent(matches[0])
    return {"applied_bookmark": bookmark_name}


def list_layouts():
    aprx = _current_project()
    return {"layouts": [lyt.name for lyt in aprx.listLayouts()]}


def export_map_to_file(out_path, resolution=150):
    """Export the active map view to PNG/PDF/TIFF (by out_path's extension)."""
    aprx = _current_project()
    view = aprx.activeView
    if not hasattr(view, "camera"):
        raise RuntimeError(
            "The active view isn't a map view. Open/activate a map view in ArcGIS "
            "Pro first, or use export_layout_to_file for a layout instead."
        )
    ext = out_path.rsplit(".", 1)[-1].lower()
    exporters = {
        "png": view.exportToPNG,
        "pdf": view.exportToPDF,
        "tif": view.exportToTIFF,
        "tiff": view.exportToTIFF,
        "jpg": view.exportToJPEG,
        "jpeg": view.exportToJPEG,
    }
    exporter = exporters.get(ext)
    if exporter is None:
        raise ValueError(f"Unsupported export extension '.{ext}'. Use png, pdf, tif, or jpg.")
    exporter(out_path, resolution=resolution) if ext != "pdf" else exporter(out_path)
    return {"exported": out_path}


def export_layout_to_file(layout_name, out_path, resolution=150):
    aprx = _current_project()
    lyt = _find_layout(aprx, layout_name)
    ext = out_path.rsplit(".", 1)[-1].lower()
    exporters = {
        "pdf": lambda p: lyt.exportToPDF(p),
        "png": lambda p: lyt.exportToPNG(p, resolution=resolution),
        "tif": lambda p: lyt.exportToTIFF(p, resolution=resolution),
        "tiff": lambda p: lyt.exportToTIFF(p, resolution=resolution),
        "jpg": lambda p: lyt.exportToJPEG(p, resolution=resolution),
        "jpeg": lambda p: lyt.exportToJPEG(p, resolution=resolution),
    }
    exporter = exporters.get(ext)
    if exporter is None:
        raise ValueError(f"Unsupported export extension '.{ext}'. Use pdf, png, tif, or jpg.")
    exporter(out_path)
    return {"exported": out_path}


def save_project():
    aprx = _current_project()
    aprx.save()
    return {"saved": aprx.filePath}


# ---------------------------------------------------------------------------
# "Reliability Inspection Form" V6.1 importer -- a form-specific bulk loader.
#
# Layout this expects (confirmed against a real V6.1 file):
#   Row 2 / Row 4: one-time per-job header info (Project ID, date, region,
#       feeder, substation, inspector, device, cost estimates).
#   Row 7/8: field labels / format hints (not used at runtime -- documented
#       here for whoever maintains this).
#   Row 9+: one row per pole/structure, ending at the first row with a blank
#       ITEM# (column A).
#
# Column C ("DLOC# / GPS coordinates") holds "(lat, lon)" as plain text.
# Rows without it are skipped -- there's no way to place them on the map.
# The labor/cost-estimating "Work Totals" columns (AQ onward) are
# intentionally NOT imported; this only carries pole condition data.
# ---------------------------------------------------------------------------

_JOB_HEADER_CELLS = {
    "ProjectID": (2, "E"),
    "InspectionDate": (2, "L"),
    "Jurisdiction": (2, "R"),
    "Region": (2, "X"),
    "Network": (2, "AE"),
    "Substation": (2, "AK"),
    "Inspector": (2, "AO"),
    "ReconductorEstCost": (4, "M"),
    "RelocationEstCost": (4, "V"),
    "SectionalizationEstCost": (4, "AD"),
    "Feeder": (4, "AH"),
    "DeviceID": (4, "AK"),
    "DeviceType": (4, "AO"),
}
_JOB_HEADER_FIELDS = [
    ("ProjectID", "TEXT", 100),
    ("InspectionDate", "DATE", None),
    ("Jurisdiction", "TEXT", 100),
    ("Region", "TEXT", 100),
    ("Network", "TEXT", 100),
    ("Substation", "TEXT", 100),
    ("Inspector", "TEXT", 100),
    ("ReconductorEstCost", "DOUBLE", None),
    ("RelocationEstCost", "DOUBLE", None),
    ("SectionalizationEstCost", "DOUBLE", None),
    ("Feeder", "TEXT", 50),
    ("DeviceID", "TEXT", 50),
    ("DeviceType", "TEXT", 50),
    ("SourceFile", "TEXT", 255),
]

# (field_name, arcpy field type, text length or None, source column letter)
_POLE_FIELDS = [
    ("ProjectID", "TEXT", 100, None),  # foreign key to Inspection_Jobs, filled in from the header
    ("ItemNum", "LONG", None, "A"),
    ("StructureBIL_kV", "DOUBLE", None, "B"),
    ("GPS_Raw", "TEXT", 60, "C"),
    ("PicRef", "TEXT", 50, "D"),
    ("TruckAccess", "TEXT", 5, "E"),
    ("TrafficControl", "TEXT", 5, "F"),
    ("CoverUpRequired", "TEXT", 5, "G"),
    ("PoleType", "TEXT", 5, "H"),
    ("WireSizeGE336", "TEXT", 5, "I"),
    ("StructureType", "TEXT", 10, "J"),
    ("BadPoleQty", "SHORT", None, "K"),
    ("BadPoleLocation", "TEXT", 5, "L"),
    ("BadCrossarmQty", "SHORT", None, "M"),
    ("BadCrossarmBraceQty", "SHORT", None, "N"),
    ("FiberglassStandoffQty", "SHORT", None, "O"),
    ("DamagedInsulatorsQty", "SHORT", None, "P"),
    ("LooseGuysQty", "SHORT", None, "Q"),
    ("BadAnchorsQty", "SHORT", None, "R"),
    ("GuyStrainInsulatorQty", "SHORT", None, "S"),
    ("ArresterQty", "SHORT", None, "T"),
    ("ArresterAction", "TEXT", 10, "U"),
    ("FuseSwitchQty", "SHORT", None, "V"),
    ("FusePhase", "TEXT", 10, "W"),
    ("RemoveGroundsQty", "SHORT", None, "X"),
    ("InstallHendrixGroundQty", "SHORT", None, "Y"),
    ("MissingPoleGroundQty", "SHORT", None, "Z"),
    ("UnfusedLateralQty", "SHORT", None, "AA"),
    ("AnimalGuardQty", "SHORT", None, "AB"),
    ("SlackConductorQty", "SHORT", None, "AC"),
    ("MissingNeutralSpans", "SHORT", None, "AD"),
    ("ConductorDamageQty", "SHORT", None, "AE"),
    ("AAACSleeveQty", "SHORT", None, "AF"),
    ("DisconnectSwitchDamageQty", "SHORT", None, "AG"),
    ("GOABSwitchDamageQty", "SHORT", None, "AH"),
    ("VegetationIssuesQty", "SHORT", None, "AI"),
    ("OtherIssuesQty", "SHORT", None, "AJ"),
    ("OtherIssuesCost", "DOUBLE", None, "AK"),
    ("WorkType", "TEXT", 10, "AL"),
    ("ReliabilityOverride", "TEXT", 20, "AM"),
    ("ImprovementOptions", "TEXT", 100, "AN"),
    ("Comments", "TEXT", 500, "AO"),
]

_GPS_PATTERN = re.compile(r"\(?\s*(-?\d+\.?\d*)\s*,\s*(-?\d+\.?\d*)\s*\)?")


def _ensure_table(path, name, fields, geometry_type=None, spatial_reference=None):
    full_path = f"{path}\\{name}"
    if arcpy.Exists(full_path):
        return full_path
    if geometry_type:
        arcpy.management.CreateFeatureclass(
            path, name, geometry_type, spatial_reference=spatial_reference
        )
    else:
        arcpy.management.CreateTable(path, name)
    for field_name, field_type, length, *_ in fields:
        kwargs = {"field_length": length} if length else {}
        arcpy.management.AddField(full_path, field_name, field_type, **kwargs)
    return full_path


def import_reliability_form(xlsx_path, sheet_name="Reliability Form"):
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError(
            "The 'openpyxl' package is not installed in ArcGIS Pro's Python "
            "environment. Open the Python Command Prompt (for your cloned "
            "environment) and run: pip install openpyxl"
        ) from exc

    aprx, m = _current_map()
    gdb = aprx.defaultGeodatabase

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Sheets in this file: {wb.sheetnames}")
    ws = wb[sheet_name]

    def cell(row, col_letter):
        return ws[f"{col_letter}{row}"].value

    header = {name: cell(row, col) for name, (row, col) in _JOB_HEADER_CELLS.items()}
    project_id = header.get("ProjectID")
    if not project_id:
        raise ValueError("Could not find a Project ID in this form (expected at E2) -- is this a V6.1 Reliability Inspection Form?")

    jobs_path = _ensure_table(gdb, "Inspection_Jobs", _JOB_HEADER_FIELDS)
    poles_path = _ensure_table(
        gdb, "Pole_Inspections", _POLE_FIELDS,
        geometry_type="POINT", spatial_reference=arcpy.SpatialReference(4326),
    )

    job_field_names = [f[0] for f in _JOB_HEADER_FIELDS]
    job_values = [xlsx_path if fn == "SourceFile" else header.get(fn) for fn in job_field_names]
    with arcpy.da.InsertCursor(jobs_path, job_field_names) as cursor:
        cursor.insertRow(job_values)

    pole_field_names = [f[0] for f in _POLE_FIELDS if f[3]]  # excludes ProjectID (no source column)
    imported, skipped_no_gps = 0, 0
    row_num = 9
    while True:
        item_num = cell(row_num, "A")
        if item_num is None:
            break
        gps_raw = cell(row_num, "C")
        match = _GPS_PATTERN.search(str(gps_raw)) if gps_raw else None
        if not match:
            skipped_no_gps += 1
            row_num += 1
            continue

        lat, lon = float(match.group(1)), float(match.group(2))
        values = [cell(row_num, col) for _, _, _, col in _POLE_FIELDS if col]
        with arcpy.da.InsertCursor(poles_path, ["ProjectID"] + pole_field_names + ["SHAPE@XY"]) as cursor:
            cursor.insertRow([project_id] + values + [(lon, lat)])
        imported += 1
        row_num += 1

    m.addDataFromPath(jobs_path)
    m.addDataFromPath(poles_path)

    return {
        "project_id": project_id,
        "poles_imported": imported,
        "poles_skipped_no_gps": skipped_no_gps,
        "jobs_table": jobs_path,
        "poles_feature_class": poles_path,
    }
